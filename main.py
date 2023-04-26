from openpyxl import Workbook,load_workbook
import os
wb = load_workbook("gorsel-isimlendirme.xlsx")
ws = wb.active
# Mr.Cloak
dizin = './gorseller/'

dosyalar = os.listdir(dizin)
excel = []
resim = []
# excel sütunları seçin
for satir in ws["B1":"B{a}".format(a=str(ws.max_row))]:
    for hucre in satir:
        excel.append(hucre.value+".jpeg")
# Dizin sütunları
for satir in ws["A1":"A{a}".format(a=str(ws.max_row))]:
    for hucre in satir:
        resim.append(hucre.value+".jpeg")
#  Kontrol şeması
def control(a):
    if a in resim:
        return True
    else:
        return False
    
for i,dosya in enumerate(dosyalar, start=0):
    if control(dosya):
        os.rename(dizin+resim[i],dizin+excel[i])
        print("{eski} olan dosya, {yeni} oldu".format(eski=resim[i],yeni=excel[i]))
    else:
        # Bu hatayı alıyorsanız klasörün içinde aradığınız dosya yoktur.
        print("Sütun1 hatalı")